# Export Service for QualityStudio
# Generates PDF and Excel reports

import io
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PDFExporter:
    """Generate PDF reports"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.grey
        ))
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold',
            textColor=colors.white
        ))
    
    def create_defects_report(
        self,
        defects: List[Dict[str, Any]],
        title: str = "Defect Report",
        subtitle: Optional[str] = None
    ) -> bytes:
        """Generate a PDF report of defects"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        # Title
        story.append(Paragraph(title, self.styles['CustomTitle']))
        
        # Subtitle with date
        if subtitle:
            story.append(Paragraph(subtitle, self.styles['CustomSubtitle']))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['CustomSubtitle']
        ))
        story.append(Spacer(1, 20))
        
        # Summary stats
        total = len(defects)
        critical = len([d for d in defects if d.get('severity') == 'critical'])
        major = len([d for d in defects if d.get('severity') == 'major'])
        minor = len([d for d in defects if d.get('severity') == 'minor'])
        
        summary_data = [
            ['Total Defects', 'Critical', 'Major', 'Minor'],
            [str(total), str(critical), str(major), str(minor)]
        ]
        
        summary_table = Table(summary_data, colWidths=[1.5*inch]*4)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb'))
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Defects table
        if defects:
            story.append(Paragraph("Defect Details", self.styles['Heading2']))
            story.append(Spacer(1, 10))
            
            headers = ['Ticket ID', 'Type', 'Line', 'Severity', 'Status', 'Date']
            data = [headers]
            
            for defect in defects[:50]:  # Limit to 50 for PDF
                data.append([
                    str(defect.get('ticketId', 'N/A'))[:15],
                    str(defect.get('defectType', 'N/A'))[:15],
                    str(defect.get('line', 'N/A'))[:10],
                    str(defect.get('severity', 'N/A'))[:10],
                    str(defect.get('status', 'N/A'))[:10],
                    str(defect.get('dateTime', ''))[:10]
                ])
            
            table = Table(data, colWidths=[1.1*inch, 1.1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.9*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb'))
            ]))
            story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_complaints_report(
        self,
        complaints: List[Dict[str, Any]],
        title: str = "Customer Complaints Report"
    ) -> bytes:
        """Generate a PDF report of customer complaints"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        # Title
        story.append(Paragraph(title, self.styles['CustomTitle']))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['CustomSubtitle']
        ))
        story.append(Spacer(1, 20))
        
        # Summary
        total = len(complaints)
        pending = len([c for c in complaints if c.get('status') == 'pending_qfir'])
        in_progress = len([c for c in complaints if 'progress' in str(c.get('status', '')).lower()])
        
        summary_data = [
            ['Total', 'Pending QFIR', 'In Progress'],
            [str(total), str(pending), str(in_progress)]
        ]
        
        summary_table = Table(summary_data, colWidths=[1.5*inch]*3)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eff6ff')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#bfdbfe'))
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Complaints table
        if complaints:
            story.append(Paragraph("Complaint Details", self.styles['Heading2']))
            story.append(Spacer(1, 10))
            
            headers = ['Ticket #', 'Customer', 'Product', 'Severity', 'Status']
            data = [headers]
            
            for complaint in complaints[:50]:
                data.append([
                    str(complaint.get('ticketNumber', 'N/A'))[:15],
                    str(complaint.get('customerName', 'N/A'))[:20],
                    str(complaint.get('productType', 'N/A'))[:15],
                    str(complaint.get('severity', 'N/A'))[:10],
                    str(complaint.get('status', 'N/A'))[:15]
                ])
            
            table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1.2*inch, 0.9*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bfdbfe'))
            ]))
            story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_kpi_report(
        self,
        kpis: List[Dict[str, Any]],
        title: str = "Quality KPI Report"
    ) -> bytes:
        """Generate a PDF report of KPIs"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        # Title
        story.append(Paragraph(title, self.styles['CustomTitle']))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['CustomSubtitle']
        ))
        story.append(Spacer(1, 20))
        
        # KPI table
        if kpis:
            headers = ['Date', 'Cpk', 'FPY %', 'Defect PPM', 'CAPA On-Time %', 'Scrap Rate %']
            data = [headers]
            
            for kpi in kpis[:30]:
                data.append([
                    str(kpi.get('recordDate', ''))[:10],
                    f"{kpi.get('cpk', 0):.2f}" if kpi.get('cpk') else 'N/A',
                    f"{kpi.get('firstPassYield', 0):.1f}" if kpi.get('firstPassYield') else 'N/A',
                    f"{kpi.get('defectPPM', 0):.0f}" if kpi.get('defectPPM') else 'N/A',
                    f"{kpi.get('onTimeCAPA', 0):.1f}" if kpi.get('onTimeCAPA') else 'N/A',
                    f"{kpi.get('scrapRate', 0):.2f}" if kpi.get('scrapRate') else 'N/A'
                ])
            
            table = Table(data, colWidths=[1*inch]*6)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecfdf5')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#6ee7b7'))
            ]))
            story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()


class ExcelExporter:
    """Generate Excel reports"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.border = Border(
            left=Side(style='thin', color='e5e7eb'),
            right=Side(style='thin', color='e5e7eb'),
            top=Side(style='thin', color='e5e7eb'),
            bottom=Side(style='thin', color='e5e7eb')
        )
    
    def create_defects_export(self, defects: List[Dict[str, Any]]) -> bytes:
        """Export defects to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Defects"
        
        # Headers
        headers = ['Ticket ID', 'Date/Time', 'Line', 'Lane', 'Shift', 'Defect Type', 
                   'Severity', 'Status', 'Inspection Method', 'Description', 'Root Cause']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        for row, defect in enumerate(defects, 2):
            ws.cell(row=row, column=1, value=defect.get('ticketId', '')).border = self.border
            ws.cell(row=row, column=2, value=str(defect.get('dateTime', ''))[:19]).border = self.border
            ws.cell(row=row, column=3, value=defect.get('line', '')).border = self.border
            ws.cell(row=row, column=4, value=defect.get('lane', '')).border = self.border
            ws.cell(row=row, column=5, value=defect.get('shift', '')).border = self.border
            ws.cell(row=row, column=6, value=defect.get('defectType', '')).border = self.border
            ws.cell(row=row, column=7, value=defect.get('severity', '')).border = self.border
            ws.cell(row=row, column=8, value=defect.get('status', '')).border = self.border
            ws.cell(row=row, column=9, value=defect.get('inspectionMethod', '')).border = self.border
            ws.cell(row=row, column=10, value=defect.get('description', '')).border = self.border
            ws.cell(row=row, column=11, value=defect.get('rootCause', '')).border = self.border
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_complaints_export(self, complaints: List[Dict[str, Any]]) -> bytes:
        """Export complaints to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Complaints"
        
        # Headers
        headers = ['Ticket Number', 'Date Logged', 'Customer', 'Product Type', 
                   'Severity', 'Status', 'Description', 'Assigned To', 'QFIR Completed']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        for row, complaint in enumerate(complaints, 2):
            ws.cell(row=row, column=1, value=complaint.get('ticketNumber', '')).border = self.border
            ws.cell(row=row, column=2, value=str(complaint.get('dateLogged', ''))[:10]).border = self.border
            ws.cell(row=row, column=3, value=complaint.get('customerName', '')).border = self.border
            ws.cell(row=row, column=4, value=complaint.get('productType', '')).border = self.border
            ws.cell(row=row, column=5, value=complaint.get('severity', '')).border = self.border
            ws.cell(row=row, column=6, value=complaint.get('status', '')).border = self.border
            ws.cell(row=row, column=7, value=complaint.get('complaintDescription', '')).border = self.border
            ws.cell(row=row, column=8, value=complaint.get('assignedTo', '')).border = self.border
            ws.cell(row=row, column=9, value='Yes' if complaint.get('qfirCompleted') else 'No').border = self.border
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_kpi_export(self, kpis: List[Dict[str, Any]]) -> bytes:
        """Export KPIs to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KPIs"
        
        # Headers
        headers = ['Date', 'Cpk', 'First Pass Yield %', 'Defect PPM', 
                   'On-Time CAPA %', 'Scrap Rate %', 'Customer Complaints']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        for row, kpi in enumerate(kpis, 2):
            ws.cell(row=row, column=1, value=str(kpi.get('recordDate', ''))[:10]).border = self.border
            ws.cell(row=row, column=2, value=kpi.get('cpk')).border = self.border
            ws.cell(row=row, column=3, value=kpi.get('firstPassYield')).border = self.border
            ws.cell(row=row, column=4, value=kpi.get('defectPPM')).border = self.border
            ws.cell(row=row, column=5, value=kpi.get('onTimeCAPA')).border = self.border
            ws.cell(row=row, column=6, value=kpi.get('scrapRate')).border = self.border
            ws.cell(row=row, column=7, value=kpi.get('customerComplaints', 0)).border = self.border
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def create_full_export(
        self,
        defects: List[Dict[str, Any]],
        complaints: List[Dict[str, Any]],
        rcas: List[Dict[str, Any]],
        capas: List[Dict[str, Any]],
        kpis: List[Dict[str, Any]]
    ) -> bytes:
        """Create a full Excel workbook with all data"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets for each entity
        self._add_defects_sheet(wb, defects)
        self._add_complaints_sheet(wb, complaints)
        self._add_rcas_sheet(wb, rcas)
        self._add_capas_sheet(wb, capas)
        self._add_kpis_sheet(wb, kpis)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _add_defects_sheet(self, wb, defects):
        ws = wb.create_sheet("Defects")
        headers = ['Ticket ID', 'Date', 'Line', 'Defect Type', 'Severity', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
        for row, defect in enumerate(defects, 2):
            ws.cell(row=row, column=1, value=defect.get('ticketId', ''))
            ws.cell(row=row, column=2, value=str(defect.get('dateTime', ''))[:10])
            ws.cell(row=row, column=3, value=defect.get('line', ''))
            ws.cell(row=row, column=4, value=defect.get('defectType', ''))
            ws.cell(row=row, column=5, value=defect.get('severity', ''))
            ws.cell(row=row, column=6, value=defect.get('status', ''))
    
    def _add_complaints_sheet(self, wb, complaints):
        ws = wb.create_sheet("Complaints")
        headers = ['Ticket #', 'Customer', 'Product', 'Severity', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
            cell.font = self.header_font
        for row, c in enumerate(complaints, 2):
            ws.cell(row=row, column=1, value=c.get('ticketNumber', ''))
            ws.cell(row=row, column=2, value=c.get('customerName', ''))
            ws.cell(row=row, column=3, value=c.get('productType', ''))
            ws.cell(row=row, column=4, value=c.get('severity', ''))
            ws.cell(row=row, column=5, value=c.get('status', ''))
    
    def _add_rcas_sheet(self, wb, rcas):
        ws = wb.create_sheet("RCAs")
        headers = ['Defect ID', 'Analysis Type', 'Root Cause', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
            cell.font = self.header_font
        for row, r in enumerate(rcas, 2):
            ws.cell(row=row, column=1, value=r.get('defectTicketId', ''))
            ws.cell(row=row, column=2, value=r.get('analysisType', ''))
            ws.cell(row=row, column=3, value=r.get('rootCause', ''))
            ws.cell(row=row, column=4, value=r.get('status', ''))
    
    def _add_capas_sheet(self, wb, capas):
        ws = wb.create_sheet("CAPAs")
        headers = ['Defect ID', 'RCA ID', 'Approval State']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
            cell.font = self.header_font
        for row, c in enumerate(capas, 2):
            ws.cell(row=row, column=1, value=c.get('defectTicketId', ''))
            ws.cell(row=row, column=2, value=c.get('rcaRecordId', ''))
            ws.cell(row=row, column=3, value=c.get('approvalState', ''))
    
    def _add_kpis_sheet(self, wb, kpis):
        ws = wb.create_sheet("KPIs")
        headers = ['Date', 'Cpk', 'FPY %', 'Defect PPM', 'CAPA On-Time %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            cell.font = self.header_font
        for row, k in enumerate(kpis, 2):
            ws.cell(row=row, column=1, value=str(k.get('recordDate', ''))[:10])
            ws.cell(row=row, column=2, value=k.get('cpk'))
            ws.cell(row=row, column=3, value=k.get('firstPassYield'))
            ws.cell(row=row, column=4, value=k.get('defectPPM'))
            ws.cell(row=row, column=5, value=k.get('onTimeCAPA'))


# Global instances
pdf_exporter = PDFExporter()
excel_exporter = ExcelExporter()
